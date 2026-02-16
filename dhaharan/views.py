from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import (
    JenisKegiatan, StatusKegiatan, Kegiatan, FotoKegiatan, Volunteer,
    Resep, BahanResep, StepsResep, TipsResep, NutrisiResep, FotoResep,
    TipeTransaksi, Transaksi, Pengurus
)
import openpyxl
from django.http import HttpResponse
from .serializers import (
    JenisKegiatanSerializer, StatusKegiatanSerializer,
    KegiatanSerializer, KegiatanListSerializer, FotoKegiatanSerializer,
    VolunteerSerializer, ResepSerializer, ResepListSerializer,
    BahanResepSerializer, StepsResepSerializer, TipsResepSerializer,
    NutrisiResepSerializer, FotoResepSerializer,
    TipeTransaksiSerializer, TransaksiSerializer,
    PengurusSerializer
)


class JenisKegiatanViewSet(viewsets.ModelViewSet):
    queryset = JenisKegiatan.objects.all()
    serializer_class = JenisKegiatanSerializer


class StatusKegiatanViewSet(viewsets.ModelViewSet):
    queryset = StatusKegiatan.objects.all()
    serializer_class = StatusKegiatanSerializer


class KegiatanViewSet(viewsets.ModelViewSet):
    queryset = Kegiatan.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'list':
            return KegiatanListSerializer
        return KegiatanSerializer
    
    def update(self, request, *args, **kwargs):
        """
        Override update to handle foto deletion before updating
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Check if there's a request to delete old photos
        # This can be done by checking if new photos are being uploaded
        # or if there's a specific flag in the request
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        
        return Response(serializer.data)
    
    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def delete_photos(self, request, pk=None):
        """
        Delete specific photos from kegiatan
        Expects: {"photo_ids": [1, 2, 3]}
        """
        kegiatan = self.get_object()
        photo_ids = request.data.get('photo_ids', [])
        
        if not photo_ids:
            return Response(
                {'error': 'photo_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = 0
        errors = []
        
        for photo_id in photo_ids:
            try:
                photo = FotoKegiatan.objects.get(id=photo_id, kegiatan=kegiatan)
                
                # Delete file from S3
                if photo.file_path:
                    try:
                        storage = photo.file_path.storage
                        if storage.exists(photo.file_path.name):
                            storage.delete(photo.file_path.name)
                    except Exception as e:
                        errors.append(f"Error deleting file for photo {photo_id}: {str(e)}")
                
                # Delete from database
                photo.delete()
                deleted_count += 1
                
            except FotoKegiatan.DoesNotExist:
                errors.append(f"Photo with id {photo_id} not found or doesn't belong to this kegiatan")
        
        response_data = {
            'deleted_count': deleted_count,
            'total_requested': len(photo_ids)
        }
        
        if errors:
            response_data['errors'] = errors
        
        return Response(response_data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def replace_all_photos(self, request, pk=None):
        """
        Delete all existing photos and prepare for new uploads
        This should be called BEFORE uploading new photos
        """
        kegiatan = self.get_object()
        
        # Get all photos for this kegiatan
        photos = FotoKegiatan.objects.filter(kegiatan=kegiatan)
        deleted_count = 0
        errors = []
        
        for photo in photos:
            try:
                # Delete file from S3
                if photo.file_path:
                    try:
                        storage = photo.file_path.storage
                        if storage.exists(photo.file_path.name):
                            storage.delete(photo.file_path.name)
                    except Exception as e:
                        errors.append(f"Error deleting file {photo.file_name}: {str(e)}")
                
                # Delete from database
                photo.delete()
                deleted_count += 1
                
            except Exception as e:
                errors.append(f"Error deleting photo {photo.id}: {str(e)}")
        
        response_data = {
            'message': f'Deleted {deleted_count} photos',
            'deleted_count': deleted_count
        }
        
        if errors:
            response_data['errors'] = errors
        
        return Response(response_data, status=status.HTTP_200_OK)


class FotoKegiatanViewSet(viewsets.ModelViewSet):
    queryset = FotoKegiatan.objects.all()
    serializer_class = FotoKegiatanSerializer
    
    def perform_create(self, serializer):
        kegiatan = serializer.validated_data.get('kegiatan')
        if kegiatan:
            import re
            # Sanitize kegiatan nama
            # Replace spaces with underscore and remove non-alphanumeric chars
            safe_nama = re.sub(r'[^a-zA-Z0-9]', '_', kegiatan.nama)
            
            # Get file from validated data
            file_obj = serializer.validated_data.get('file_path')
            if file_obj:
                # Sanitize original filename
                safe_filename = re.sub(r'[^a-zA-Z0-9.]', '_', file_obj.name)
                # Set new name
                file_obj.name = f"{safe_nama}_{safe_filename}"
        
        serializer.save()

    
    def destroy(self, request, *args, **kwargs):
        """
        Override destroy to delete file from S3 before deleting from database
        """
        instance = self.get_object()
        
        # Delete file from S3 if exists
        if instance.file_path:
            try:
                # Get storage instance and delete from S3
                storage = instance.file_path.storage
                if storage.exists(instance.file_path.name):
                    storage.delete(instance.file_path.name)
            except Exception as e:
                # Log error but continue with database deletion
                print(f"Error deleting file from S3: {str(e)}")
        
        # Delete from database
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class VolunteerViewSet(viewsets.ModelViewSet):
    queryset = Volunteer.objects.all()
    serializer_class = VolunteerSerializer
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        volunteer = self.get_object()
        volunteer.is_approved = True
        volunteer.save()
        serializer = self.get_serializer(volunteer)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        pending_volunteers = Volunteer.objects.filter(is_approved=False)
        serializer = self.get_serializer(pending_volunteers, many=True)
        return Response(serializer.data)


class ResepViewSet(viewsets.ModelViewSet):
    queryset = Resep.objects.all()
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ResepListSerializer
        # For retrieve, create, update, partial_update use full serializer
        return ResepSerializer
    
    @action(detail=False, methods=['get'])
    def by_kategori(self, request):
        kategori = request.query_params.get('kategori', None)
        if kategori:
            resep = Resep.objects.filter(kategori=kategori)
            serializer = self.get_serializer(resep, many=True)
            return Response(serializer.data)
        return Response(
            {'error': 'Parameter kategori diperlukan'},
            status=status.HTTP_400_BAD_REQUEST
        )


class BahanResepViewSet(viewsets.ModelViewSet):
    queryset = BahanResep.objects.all()
    serializer_class = BahanResepSerializer


class StepsResepViewSet(viewsets.ModelViewSet):
    queryset = StepsResep.objects.all()
    serializer_class = StepsResepSerializer


class TipsResepViewSet(viewsets.ModelViewSet):
    queryset = TipsResep.objects.all()
    serializer_class = TipsResepSerializer


class NutrisiResepViewSet(viewsets.ModelViewSet):
    queryset = NutrisiResep.objects.all()
    serializer_class = NutrisiResepSerializer


class FotoResepViewSet(viewsets.ModelViewSet):
    queryset = FotoResep.objects.all()
    serializer_class = FotoResepSerializer
    
    def destroy(self, request, *args, **kwargs):
        """
        Override destroy to delete file from S3 before deleting from database
        """
        instance = self.get_object()
        
        # Delete file from S3 if exists
        if instance.file_path:
            try:
                # file_path is now a URL string, extract the S3 key
                # URL format: https://s3.region.amazonaws.com/bucket/prefix/folder/file.jpg
                # We need to delete using boto3
                import boto3
                from django.conf import settings
                
                s3_client = boto3.client(
                    's3',
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME
                )
                
                # Extract key from URL
                # Example URL: https://s3.ap-southeast-1.amazonaws.com/bucket/dhaharan/resep/file.jpg
                # Extract: dhaharan/resep/file.jpg
                url_parts = instance.file_path.split('/')
                # Find index of bucket name and get everything after it
                bucket_name = settings.AWS_STORAGE_BUCKET_NAME
                bucket_index = url_parts.index(bucket_name)
                s3_key = '/'.join(url_parts[bucket_index + 1:])
                
                s3_client.delete_object(Bucket=bucket_name, Key=s3_key)
                print(f"Deleted file from S3: {s3_key}")
            except Exception as e:
                # Log error but continue with database deletion
                print(f"Error deleting file from S3: {str(e)}")
        
        # Delete from database
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class TipeTransaksiViewSet(viewsets.ModelViewSet):
    queryset = TipeTransaksi.objects.all()
    serializer_class = TipeTransaksiSerializer


class TransaksiViewSet(viewsets.ModelViewSet):
    queryset = Transaksi.objects.all()
    serializer_class = TransaksiSerializer
    
    @action(detail=False, methods=['get'])
    def by_tipe(self, request):
        tipe = request.query_params.get('tipe', None)
        if tipe:
            transaksi = Transaksi.objects.filter(tipe_transaksi_id=tipe)
            serializer = self.get_serializer(transaksi, many=True)
            return Response(serializer.data)
        return Response(
            {'error': 'Parameter tipe diperlukan'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        from django.db.models import Sum
        total_pemasukan = Transaksi.objects.filter(
            tipe_transaksi__nama='Pemasukan'
        ).aggregate(total=Sum('jumlah'))['total'] or 0
        
        total_pengeluaran = Transaksi.objects.filter(
            tipe_transaksi__nama='Pengeluaran'
        ).aggregate(total=Sum('jumlah'))['total'] or 0
        
        return Response({
            'total_pemasukan': total_pemasukan,
            'total_pengeluaran': total_pengeluaran,
            'saldo': total_pemasukan - total_pengeluaran
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        group_by = request.query_params.get('group_by', 'monthly')  # monthly, yearly
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=cashflow_report_{group_by}.xlsx'
        
        workbook = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        queryset = self.filter_queryset(self.get_queryset()).order_by('-tanggal')
        
        # Group data
        grouped_data = {}
        
        # Helper for Indonesian month names
        bulan_indo = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        
        for item in queryset:
            date = item.tanggal
            if group_by == 'yearly':
                key = str(date.year)
            else: # monthly
                key = f"{bulan_indo[date.month]} {date.year}"
            
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(item)
            
        # Create sheets
        # Sort keys to have order (Newest first seems appropriate as per UI, but Excel usually left-to-right. 
        # Let's do Newest first for relevance)
        sorted_keys = sorted(grouped_data.keys(), key=lambda x: parse_group_key(x, group_by), reverse=True)

        if not sorted_keys:
             workbook.create_sheet("No Data")

        for key in sorted_keys:
            items = grouped_data[key]
            # Sanitize sheet title (max 31 chars, no invalid chars)
            safe_title = key[:31].replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
            worksheet = workbook.create_sheet(title=safe_title)
            
            # Header
            headers = ['No', 'Nama Transaksi', 'Deskripsi', 'Tipe', 'Jumlah', 'Tanggal']
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                
            # Data
            for row_num, item in enumerate(items, 2):
                worksheet.cell(row=row_num, column=1, value=row_num-1)
                worksheet.cell(row=row_num, column=2, value=item.nama)
                worksheet.cell(row=row_num, column=3, value=item.deskripsi)
                worksheet.cell(row=row_num, column=4, value=item.tipe_transaksi.nama if item.tipe_transaksi else '-')
                worksheet.cell(row=row_num, column=5, value=item.jumlah)
                worksheet.cell(row=row_num, column=6, value=str(item.tanggal))

            # Auto-adjust column width (simple approx)
            for col in worksheet.columns:
                 max_length = 0
                 column = col[0].column_letter # Get the column name
                 for cell in col:
                     try:
                         if len(str(cell.value)) > max_length:
                             max_length = len(str(cell.value))
                     except:
                         pass
                 adjusted_width = (max_length + 2)
                 worksheet.column_dimensions[column].width = adjusted_width
            
        workbook.save(response)
        return response

def parse_group_key(key, group_by):
    # Helper to sort keys correctly since they are strings
    if group_by == 'yearly':
        return int(key)
    else:
        # Parse "Januari 2024" back to date object for sorting
        try:
            parts = key.split(' ')
            bulan_map = {
                'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4, 'Mei': 5, 'Juni': 6,
                'Juli': 7, 'Agustus': 8, 'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
            }
            month = bulan_map.get(parts[0], 1)
            year = int(parts[1])
            from datetime import date
            return date(year, month, 1)
        except:
             return 0


class PengurusViewSet(viewsets.ModelViewSet):
    queryset = Pengurus.objects.all().order_by('id')
    serializer_class = PengurusSerializer



